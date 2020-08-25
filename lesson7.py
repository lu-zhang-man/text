# -*- coding: utf-8 -*-
# @Author   :   YaMeng
# @File :   lesson7.py
# @Software :   PyCharm
# @Time :   2020/8/21 14:05
# @company  :   湖南省零檬信息技术有限公司

# 接口自动化的步骤
# 使用excel整理好接口测试用例   -- done
# 通过python去读取excel的测试数据  -- done
# 使用requests去发送请求，并且得到响应结果  -- done
# 断言：执行结果  vs  预期结果   -- done
# 把通过/不通过的最终结果要回写到excel   -- done

import requests
import openpyxl


# 读取excel测试用例
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row  # 获取sheet里最大的行数
    case_list = []    # 定义一个空列表，是专门用来接收测试数据的
    for i in range(2, max_row+1):  # 因为取头不取尾，所以要在最大行数的基础上 +1
        dict1 = dict(   #  转换成字典就是为了将测试用例打包成一条一条的数据
        id = sheet.cell(row=i, column=1).value,        #  取的是测试用来编号id
        url = sheet.cell(row=i, column=5).value,       #  取的是url数据
        data = sheet.cell(row=i, column=6).value,      #  取的是data数据
        expect = sheet.cell(row=i, column=7).value)    #  取的是预期结果（expected数据）
        case_list.append(dict1)     # 用append把字典追加到列表去  --> 列表就存放了所有的测试数据
    # print(case_list)
    return case_list   # 设置返回值，给别人去用

# 发送请求
def api_func(url, res_body):
    requests_header = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
    res = requests.post(url=url, json=res_body, headers=requests_header)
    res_log = res.json()
    return res_log

# 把测试结果写入到excel
def writr_result(filename, sheetname, row, colummn, final_result):
    wb = openpyxl.load_workbook(filename)  # 加载工作薄
    sheet = wb[sheetname]
    sheet.cell(row=row, column=colummn).value = final_result  # 加入最终结果
    wb.save(filename)  # 保存excel

# 执行接口测试，并断言。回写测试结果到excel
def exeture_func(filename, sheetname):
    res = read_data(filename, sheetname)
    for testcase in res:     # 取出一条一条的测试用例
        case_id = testcase.get('id')    # 字典取值或者value，取出id
        url = testcase.get('url')     # 取出url
        data = testcase.get('data')    # 取出 data 从excel取出来数据都是str
        data = eval(data)  # 运行被字符串包裹的python表达式，转换成字典格式    -- 用eval()把引号去掉
        expect = testcase.get('expect')  # 取出expect
        expect = eval(expect)   # 把字符串转换成字典
        expect_msg = expect.get('msg')   # 从预期结果的字典里把msg取出来
        # print(case_id, url, data, expect)
        res_1 = api_func(url=url,res_body=data)   # 调用发送请求的函数，并传入参数
        # print(res_1)
        real_msg = res_1.get('msg')  # 把实际结果里的msg取出来
        # print(expect_msg, real_msg)
        print('预期结果为：{}'.format(expect_msg))
        print('实际结果为：{}'.format(real_msg))
        if real_msg == expect_msg:
            print('这条测试用例执行通过！')
            final_res = '通过'
        else:
            print('这条测试用例执行不通过！！！！')
            final_res = '不通过，有bug'
        print('*' * 100)
        writr_result(filename,sheetname,case_id+1, 8, final_res)


exeture_func('test_case_api.xlsx', 'register')
exeture_func('test_case_api.xlsx', 'login')